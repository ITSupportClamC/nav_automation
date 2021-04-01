# coding=utf-8
#
# Functions needed to calculate IMA yield
# 
import ast
import hashlib
import json
import logging
import math
import os
import sys
import time
import openpyxl
from datetime import datetime
from itertools import tee
from os.path import abspath, dirname
from nav_automation.constants import Constants

import requests
import xlrd

from nav_automation.constants import Constants


def getSTBFNavDataFromFile(file):
	"""
	[String] file (daily NAV file for Short Term Bond Fund)
		=> [Iterable] ( date (yyyy-mm-dd)
					  , class
					  , currency
					  , number of unit
					  , total nav of the class
					  , nav per unit)
	"""
	return NavHandler().getSTBFNavDataFromFile(file)
	# return []



def updateWebSite(mode, timeOut, fundName, navData):
	"""
	[String] mode (0 means production, 1 means test)
	[String] timeout (in miliseconds)
	[String] fund name,
	[Tuple] ( date (yyyy-mm-dd)
			, class
			, currency
			, number of unit
			, total nav of the class
			, nav per unit)

	In production mode (0), update the website in production.
	In test mode (1), update the test website.

	Web sites (production, test) and login credentials should be
	configurable in a file.

	If timed out when trying to update the web site, throw an exception.

	return 0 when successful, throw exception otherwise.
	"""
	return NavHandler().updateWebSite(mode, timeOut, fundName, navData)



def createBloombergExcelFile(templateFile, outputDir, fundName, data):
	"""
	[String] template file,
	[String] output directory,
	[String] fund name
	[Iterable] ( date (yyyy-mm-dd)
			   , class
			   , currency
			   , number of unit
			   , total nav of the class
			   , nav per unit)

	=> [String] output file

	Assume: the list of classes in data all belong to the same fund.

	Side effect: create an excel file in the output directory,
	with the data populated. The template file should be kept
	unchanged.
	"""
	return NavHandler().createBloombergExcelFile(templateFile, outputDir, fundName, data)



def createThomsonExcelFile(templateFile, outputDir, fundName, data):
	"""
	[String] template file,
	[String] output directory,
	[String] fund name
	[Iterable] ( date (yyyy-mm-dd)
			   , class
			   , currency
			   , number of unit
			   , total nav of the class
			   , nav per unit)

	=> [String] output file

	Assume: the list of classes in data all belong to the same fund.

	Side effect: create an excel file in the output directory,
	with the data populated. The template file should be kept
	unchanged.
	"""
	return NavHandler().createThomsonExcelFile(templateFile, outputDir, fundName, data)



def getBloombergCode(fundName, className, currency):
	"""
	[String] fund name,
	[String] class name,
	[String] currency
		=> [String] Bloomberg code
	"""
	f_map = \
	{ ('stbf', 'Class B', 'USD'): 'CLSTFBU HK Equity'
	, ('stbf', 'Class I', 'USD'): 'CLSTFIU HK Equity'
	, ('stbf', 'Class A', 'USD'): 'CLSTFAU HK Equity'
	, ('stbf', 'Class A (USD)', 'USD'): 'CLSTFAU HK Equity'
	, ('stbf', 'Class A', 'HKD'): 'CLSTFAH HK Equity'
	, ('stbf', 'Class A (HKD)', 'HKD'): 'CLSTFAH HK Equity'
	}
	return f_map[(fundName, className, currency)]



def getBloombergFundname(fundName):
	"""
	"""
	f_map = {'stbf': 'CHINA LIFE FR ST BOND'}
	return f_map[fundName]



def getThomsonReutersFundname(fundName):
	"""
	[String] fund name,
	[String] class name
		=> [Tuple] ISIN code, Thomson Reuters fund name
	"""
	f_map = {'stbf': 'China Life Franklin Global-Short Term Bond'}
	return f_map[fundName]



def getISINCode(fundName, className, currency='USD'):
	"""
	[String] fund name,
	[String] class name,
	[String] currency
		=> [String] ISIN code
	"""
	f_map = \
	{ ('stbf', 'Class B', 'USD'): 'HK0000664455'
	, ('stbf', 'Class I', 'USD'): 'HK0000664489'
	, ('stbf', 'Class A', 'USD'): 'HK0000664422'
	, ('stbf', 'Class A (USD)', 'USD'): 'HK0000664422'
	, ('stbf', 'Class A', 'HKD'): 'HK0000664430'
	, ('stbf', 'Class A (HKD)', 'HKD'): 'HK0000664430'
	}
	return f_map[(fundName, className, currency)]



"""
	returns the current directory

	for running the test case provided
"""
getCurrentDirectory = lambda : \
	dirname(abspath(__file__))

class NavHandler:

	def __init__(self):
		self.logger = logging.getLogger(__name__)

	#-- run() is used by the main application to demo the functions
	def run(self, file, mode, timeOut, fundName):
		self.logger.info("Program initiated")
		self.logger.debug("")  
		self.logger.debug("|============================NEW Session============================|")
		self.logger.debug("Reading file: '" + file + "'")
		#-- loop the iterator and send request to website
		all_nav = list(NavHandler().getSTBFNavDataFromFile(file))
		self.logger.info("NAV info retrieved")
		for i in range(len(all_nav)):
			self.logger.debug("")
			self.logger.debug("Prepare to upload NAV class number " + str(i + 1) + "...")
			self.logger.debug("Data retrieved: " + str(all_nav[i]))
			NavHandler().updateWebSite(mode, timeOut, fundName, all_nav[i])
			self.logger.debug("")
			self.logger.debug("Upload completed")
			self.logger.debug("|============================END Session============================|")
			self.logger.debug("")
		
	def getSTBFNavDataFromFile(self, file):
		# self.logger.info('Running operation on file: ' + file)

		def _create_iterator(input_list):
			for item in input_list:
				yield item
		
		# self.logger.info("Running operation on file: " + file)

		#-- open the xls file
		book = xlrd.open_workbook(file)
		worksheet = book.sheet_by_name('Report')
		number_of_rows = worksheet.nrows - 1
		number_of_cols = worksheet.ncols - 1

		#-- searching for data and sections ===========================================================
		date_pos = []
		unit_section_pos = []
		number_of_classes = 0
		class_pos_row = 0
		class_pos_col = []
		currency_pos_row = 0
		number_of_units_pos_filtered = 0
		number_of_units_pos_row = 0
		nav_per_unit_pos_row = 0
		total_nav_of_class_pos_row = 0
		required_data_check = [0, 0, 0, 0, 0, 0, 0]
		missing_data = ""
		data_name = ["Date", "Unit section", "Currency", "No. of units (Actual from UT-Pro)", "NAV after Management Fee", "NAV per Unit (in 4 dec.)", "Classes"]
		
		#-- search for data and sections
		for r in range(number_of_rows):
			for c in range(number_of_cols):
				#-- finding date
				if ((not date_pos) and "Dealing Date" in str(worksheet.cell_value(r, c))):
					date_pos.extend([r, c])
					# self.logger.info("Date found! (at row: " + str(r) + " | col: " + str(c) + ")")
					required_data_check[0] = 1
				
				#-- finding unit section
				if ((not unit_section_pos) and str(worksheet.cell_value(r, c)) == "UNITS"):
					unit_section_pos.extend([r, c])
					class_pos_row = r + 1
					# self.logger.info("Unit section found! (at row: " + str(r) + " | col: " + str(c) + ")")
					required_data_check[1] = 1

				#-- finding currency section
				if ( (unit_section_pos) and (not currency_pos_row) and ("in class currency" in str(worksheet.cell_value(r, c))) ):
					currency_pos_row = r
					# self.logger.info("Currency found! (at row: " + str(r) + " | col: " + str(c) + ")")
					required_data_check[2] = 1

				#-- finding no. of units section
				if ( number_of_units_pos_filtered and (unit_section_pos) and (not number_of_units_pos_row) and ("No of units (Actual from UT-Pro)" in str(worksheet.cell_value(r, c))) ):
					number_of_units_pos_row = r
					# self.logger.info("Number of units found! (at row: " + str(r) + " | col: " + str(c) + ")")
					required_data_check[3] = 1

				#-- ensuring that the no. of units section is the second instance of itself
				if (("No of units (Actual from UT-Pro)" in str(worksheet.cell_value(r, c)))):
					number_of_units_pos_filtered = 1

				#-- finding total nav of the class
				if ( (unit_section_pos) and (not total_nav_of_class_pos_row) and ("NAV after Management Fee" in str(worksheet.cell_value(r, c))) ):
					total_nav_of_class_pos_row = r
					# self.logger.info("Totla NAV of the class found! (at row: " + str(r) + " | col: " + str(c) + ")")
					required_data_check[4] = 1

				#-- finding nav per unit section
				if ( (unit_section_pos) and (not nav_per_unit_pos_row) and ("NAV per Unit (in 4 dec.)" in str(worksheet.cell_value(r, c))) ):
					nav_per_unit_pos_row = r
					# self.logger.info("NAV per unit found! (at row: " + str(r) + " | col: " + str(c) + ")")
					required_data_check[5] = 1

		for c in range(number_of_cols):
			if ("Class" in str(worksheet.cell_value(class_pos_row, c))):
				number_of_classes += 1
				class_pos_col.append(c)
		if (number_of_classes <= 0):
			error_message = "No class found, data retrieval failed. Please check if all classes' names are listed on the next row of 'UNITS'."
			# self.logger.error(error_message)
			raise ValueError(error_message)
		else:
			# self.logger.info(str(number_of_classes) + " class(es) found! (at row: " + str(class_pos_row) + " | col: " + str(class_pos_col) + ")")
			required_data_check[6] = 1

		if (0 in required_data_check):
			for i in range(len(required_data_check)):
				if (required_data_check[i]) == 0:
					missing_data += data_name[i] + ", "
			error_message = "Data retrieval failed, the following data/sections is(are) perhaps incorrect or missing: " + missing_data + "please double check the worksheet."
			# self.logger.error(error_message)
			raise ValueError(error_message)
		else:
			self.logger.debug("All data present, operation begins!")
			# self.logger.info("All data present, operation begins!")
		#-- ===============================================================================================

		#-- gathering data ================================================================================

		#-- retrieving dealing date
		try:
			dealing_date = str(worksheet.cell_value(date_pos[0], date_pos[1])).replace("Dealing Date: " ,"")
			datetimeobject = datetime.strptime(dealing_date,'%d %B %Y')
			dealing_date = datetimeobject.strftime('%Y-%m-%d')
		except ValueError:
			error_message = "Date format conversion failed, perhaps it is incorrect or missing? (expected format e.g.: '01 January 2021)')"
			# self.logger.error(error_message)
			raise ValueError(error_message)
		
		dealing_date_list = []
		currency = []
		class_name = []
		number_of_units = []
		total_nav_of_class =[]
		nav_per_unit = []
		data_tuple = []

		for i in range(number_of_classes):
			#-- appending date to each instance of class
			dealing_date_list.append(dealing_date)

			#-- retrieving class name
			class_name.append(str(worksheet.cell_value(class_pos_row, class_pos_col[i])))

			#-- retrieving currency
			#-- dealing with the exceptional case placement of the currency column for the first class
			if (i <= 0):
				if (str(worksheet.cell_value(currency_pos_row, class_pos_col[i]-1)) != ''):
					currency.append(str(worksheet.cell_value(currency_pos_row, class_pos_col[i]-1)))
				else:
					error_message = "Currency retrieval failed, please check if the currency field has already been filled in for every class."
					# self.logger.error(error_message)
					raise ValueError(error_message) 
			else:
				if (str(worksheet.cell_value(currency_pos_row, class_pos_col[i]-2)) != ''):
					currency.append(str(worksheet.cell_value(currency_pos_row, class_pos_col[i]-2)))
				else:
					error_message = "Currency retrieval failed, please check if the currency field has already been filled in for every class."
					# self.logger.error(error_message)
					raise ValueError(error_message)         

			#-- retrieving number of units
			if (str(worksheet.cell_value(number_of_units_pos_row, class_pos_col[i])) != ''):
				number_of_units.append(worksheet.cell_value(number_of_units_pos_row, class_pos_col[i]))
			else:
				error_message = "Number of units retrieval failed, please check if the second 'No of units (Actual from UT-Pro)' field has already been filled in for every class."
				# self.logger.error(error_message)
				raise ValueError(error_message)

			#-- retrieving total nav of the class
			if (str(worksheet.cell_value(total_nav_of_class_pos_row, class_pos_col[i])) != ''):
				total_nav_of_class.append(worksheet.cell_value(total_nav_of_class_pos_row, class_pos_col[i]))
			else:
				error_message = "Total number of NAV of the class retrieval failed, please check if the second 'NAV after Management Fee' field has already been filled in for every class."
				# self.logger.error(error_message)
				raise ValueError(error_message)   

			#-- retrieving nav per unit
			if (str(worksheet.cell_value(nav_per_unit_pos_row, class_pos_col[i])) != ''):
				nav_per_unit.append(worksheet.cell_value(nav_per_unit_pos_row, class_pos_col[i]))
			else:
				error_message = "NAV per unit retrieval failed, please check if the 'NAV per Unit (in 4 dec.)' field has already been filled in for every class."
				# self.logger.error(error_message)
				raise ValueError(error_message)

			#-- creating tuple for each class
			each_tuple = (dealing_date_list[i], class_name[i], currency[i], number_of_units[i], total_nav_of_class[i], nav_per_unit[i])
			data_tuple.append(each_tuple)
		#-- ===============================================================================================

		ouput_data = _create_iterator(data_tuple)
		# self.logger.info("Data retrieval successful! operation completed!")
		# self.logger.info("Retrieved data:")
		# self.logger.info(data_tuple)
		return ouput_data

	def updateWebSite(self, mode, timeOut, fundName, navData):
		#-- selecting between production server and test server
		if (mode == Constants.MODE_TEST):
			url = Constants.API_TEST
		elif (mode == Constants.MODE_PRODUCTION):
			url = Constants.API_PRD
		else:
			error_message = "Unknown mode: " + str(mode)
			self.logger.error(error_message)
			raise ValueError(error_message)

		#-- timeOut data type checking
		if type(timeOut) != int:
			error_message = "timeOut requires to be an intetger: " + timeOut
			self.logger.error(error_message)
			raise ValueError(error_message)			

		#-- data type checking
		productId = -1
		try:
			productId = self._getWebsiteProductID(fundName)
		except KeyError:
			error_message = "Failed to find productId from the provided fundName. " + \
							"Please create the product on website and add the mapping in function _getWebsiteProductID()."
			self.logger.error(error_message)
			raise KeyError(error_message)

		# initializing tuples
		payload_keys = ('date', 'className', 'currency', 'numOfUnits', 'totalNumOfNav', 'navPerUnit') 
		payload_values = navData

		payload = dict()
		payload["auth_token"] = str("")
		#-- using dictionary comprehension 
		#-- convert tuples to dictionary 
		if len(payload_keys) == len(payload_values): 
			nav_payload = {payload_keys[i] : payload_values[i] for i, _ in enumerate(payload_values)}
		#-- merge two dictionaries
		payload = {**payload, **nav_payload}
		payload["productId"] = productId
		payload["created_at"] = math.floor(time.time())
		#-- create description
		payload["description"] = str(payload.get("className")) + \
									"-" + \
									str(payload.get("currency")) + \
									"$ " + str(payload.get("navPerUnit"))
		#-- generate auth_token
		md5_str = str(payload.get("productId")) + \
					Constants.API_TOKEN_STR + \
					str(payload.get("created_at"))
		payload["auth_token"] = hashlib.md5(md5_str.encode('utf-8')).hexdigest()

		files=[]
		headers = {}
		resp = ""

		#-- call the website API
		self.logger.info("Calling API: " + \
							url + \
							", data: " + \
							str(payload)
							)
		
		try:
			resp = requests.request("POST", url, headers=headers, data=payload, files=files, timeout=timeOut)
			#-- check and raise error if http code is not 200 (good)
			#-- skip checking 400 as want to print the params validation result
			if (resp.status_code != 400):
				resp.raise_for_status()
			resp_content = json.loads(resp.text)
			resp_result = resp_content['meta']["result"]
			if (resp_result != "success"):
				error_message = "API response 'failed'. Reason: " + resp_content['meta']["message"]
				self.logger.error(error_message)
				raise Exception(error_message)
		except requests.exceptions.HTTPError as e:
			#-- capture all non 200 response
			error_message = "Error: " + str(e)
			self.logger.error(error_message)
			raise requests.exceptions.HTTPError(error_message)
		except ValueError as e:
			#-- capture if response cannot be parased as json
			error_message = "The response from API is not in json format: " + resp.text
			self.logger.error(error_message)
			raise ValueError(error_message)
		except Exception as e:
			#-- capture if response cannot be parased as json
			error_message = "System error when processing API. " + str(e)
			self.logger.error(error_message)
			raise Exception(error_message)

		self.logger.debug("Response from API: " + resp_result)
		
		return 0

	#-- get the productId by fundName. productId is the identifier of a fund created on the clamc site.
	def _getWebsiteProductID(self, fundName):
		"""
		[String] fund name
			=> [String] webite product ID
		"""
		f_map = \
		{ 
			'stbf': '4'
		}
		return f_map[fundName]

	def createBloombergExcelFile(self, templateFile, outputDir, fundName, data):
		wb_obj = openpyxl.load_workbook(templateFile)
		#-- assume the template is at the first active sheet
		ws_obj = wb_obj.active
		column_count = ws_obj.max_column
		column_title_row = -1

		#-- get the column title row
		for row in range(1, column_count):
			if (ws_obj.cell(row, 1).value == "DATE (MM/DD/YYYY)"):
				column_title_row = row
				break
		#-- get the first content row
		row += 1
		self.logger.debug('first data row is ' + str(row))
		#-- get the fund size actual
		total_fund_size = 0
		#-- navData format
		keys = ['date', 'className', 'currency', 'numOfUnits', 'totalNumOfNav', 'navPerUnit']
		data_l = []
		#-- convert data from tuple to dict to make value retrieval below code clearer 
		for datum in data:
			data_l.append( dict(zip(keys, list(datum))) )
		for data_d in data_l:
			total_fund_size += float(data_d['totalNumOfNav'])
		#-- convert format from yyyy-mm-dd to mm//dd/yyyy 
		def format_date(old_format):
			date_time_object = datetime.strptime(old_format,'%Y-%m-%d')
			new_format = date_time_object.strftime('%m/%d/%Y')
			return new_format
		#-- get fund name
														
		#-- write the content
		for data_d in data_l:
			ws_obj.cell(row, Constants.BG_COL_DATE).value = format_date(data_d['date'])
			try:
				ws_obj.cell(row, Constants.BG_COL_BLOOMBERG_CODE).value = getBloombergCode(fundName, 
																				data_d['className'],
																				data_d['currency'])
			except KeyError:
				error_message = "Failed to find bloombergCode from the provided " + \
								"fundName: " + fundName + ", " + \
								"className: " + data_d['className'] + ", " + \
								"currency: " + data_d['currency'] + \
								". Please add the mapping in function getBloombergCode()."
				self.logger.error(error_message)
				raise KeyError(error_message)
			#-- assume data_d['className'] have pattern of "Class X"
			try:
				ws_obj.cell(row, Constants.BG_COL_FUND_NAME).value = getBloombergFundname(fundName) + \
															'-' + \
															data_d['className'].replace('Class ', '') + \
															' ' + \
															data_d['currency']
			except KeyError:
				error_message = "Failed to find getBloombergFundname from the provided " + \
								"fundName: " + fundName + ", " + \
								". Please add the mapping in function getBloombergFundname()."
				self.logger.error(error_message)
				raise KeyError(error_message)
			ws_obj.cell(row, Constants.BG_COL_CURRENCY).value = data_d['currency']
			ws_obj.cell(row, Constants.BG_COL_NAV).value = data_d['navPerUnit']
			ws_obj.cell(row, Constants.BG_COL_BID).value = ''
			ws_obj.cell(row, Constants.BG_COL_OFFER).value = ''
			ws_obj.cell(row, Constants.BG_COL_FUND_SIZE_ACTUAL).value = total_fund_size
			ws_obj.cell(row, Constants.BG_COL_CLASS_ASSETS_ACTUAL).value = data_d['totalNumOfNav']
			ws_obj.cell(row, Constants.BG_COL_SHARE_OUT_ACTUAL).value = data_d['numOfUnits']
			ws_obj.cell(row, Constants.BG_COL_FIRM_ASSETS_UNDER_MANAGEMENT_ACTUAL).value = ''
			#increment the row count
			row += 1

		#-- save the output to the target output directory
		output_file_name = 'bloomberg_' + fundName + '_' + str(math.floor(time.time())) + ".xlsx"
		output_file_fullpath = os.path.join(outputDir, output_file_name)
		wb_obj.save(output_file_fullpath)
		return output_file_fullpath

	def createThomsonExcelFile(self, templateFile, outputDir, fundName, data):
		wb_obj = openpyxl.load_workbook(templateFile)
		#-- assume the template is at the first active sheet
		ws_obj = wb_obj.active
		column_count = ws_obj.max_column
		column_title_row = -1

		#-- get the column title row
		for row in range(1, column_count):
			if (ws_obj.cell(row, 1).value == "ISIN Code"):
				column_title_row = row
				break
		#-- get the first content row
		row += 1
		self.logger.debug('first data row is ' + str(row))
		#-- get the fund size actual
		total_fund_size = 0
		#-- navData format
		keys = ['date', 'className', 'currency', 'numOfUnits', 'totalNumOfNav', 'navPerUnit']
		data_l = []
		#-- convert data from tuple to dict to make value retrieval below code clearer 
		for datum in data:
			data_l.append( dict(zip(keys, list(datum))) )
		for data_d in data_l:
			total_fund_size += float(data_d['totalNumOfNav'])
		#-- write the content
		for data_d in data_l:
			try:
				ws_obj.cell(row, Constants.RE_COL_ISIN_CODE).value = getISINCode(fundName, 
																				data_d['className'],
																				data_d['currency'])
			except KeyError:
				error_message = "Failed to find ISIN from the provided " + \
								"fundName: " + fundName + ", " + \
								"className: " + data_d['className'] + ", " + \
								". Please add the mapping in function getISINCode()."
				self.logger.error(error_message)
				raise KeyError(error_message)
			try:
				ws_obj.cell(row, Constants.RE_COL_NAME).value = getThomsonReutersFundname(fundName) + \
															' ' + \
															data_d['className'].replace('Class ', '') + \
															' ' + \
															data_d['currency']
			except KeyError:
				error_message = "Failed to find getThomsonReutersFundname from the provided " + \
								"fundName: " + fundName + ", " + \
								". Please add the mapping in function getThomsonReutersFundname()."
				self.logger.error(error_message)
				raise KeyError(error_message)
			ws_obj.cell(row, Constants.RE_COL_CURRENCY).value = data_d['currency']
			ws_obj.cell(row, Constants.RE_COL_NAV_PER_SHARE).value = data_d['navPerUnit']
			ws_obj.cell(row, Constants.RE_COL_FUND_SIZE).value = total_fund_size
			ws_obj.cell(row, Constants.RE_COL_CLASS_ASSETS).value = data_d['totalNumOfNav']
			ws_obj.cell(row, Constants.RE_COL_SHARE_OUT).value = data_d['numOfUnits']
			#increment the row count
			row += 1

		#-- save the output to the target output directory
		output_file_name = 'reuters_' + fundName + '_' + str(math.floor(time.time())) + ".xlsx"
		output_file_fullpath = os.path.join(outputDir, output_file_name)
		wb_obj.save(output_file_fullpath)
		return output_file_fullpath
		
