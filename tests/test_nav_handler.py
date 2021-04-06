# # coding=utf-8
# # 

import unittest2
import openpyxl
import logging.config
from nav_automation.constants import Constants
from nav_automation.nav_handler import NavHandler, getCurrentDirectory
from datetime import datetime
from os.path import abspath, dirname, join

class TestNavHandler(unittest2.TestCase):

	def __init__(self, *args, **kwargs):
		super(TestNavHandler, self).__init__(*args, **kwargs)
		logging.config.fileConfig( join(getCurrentDirectory(), "logging_config.ini"),
									defaults={'date':datetime.now().date().strftime('%Y-%m-%d')}
								)

	def setUp(self):
		self.NavHandler = NavHandler()

	def testGetSTBFNavDataFromFile(self):
		file = join(getCurrentDirectory(), 'samples', 'sample PriceSTBF.xls')
		data = self.NavHandler.getSTBFNavDataFromFile(file)
		sortedData = sorted(data, key=lambda t: t[1])
		self.assertEqual(2, len(sortedData))
		self.assertEqual(
			('2020-12-30', 'Class B', 'USD', 400000, 3998441.60, 9.9961)
		  , sortedData[0]
		)
		self.assertEqual(
			('2020-12-30', 'Class I', 'USD', 4500000, 44981729.87, 9.9959)
		  , sortedData[1] #-- modified value to compare the right output (from 'sortedData[0]' to 'sortedData[1]')
		)

	def testCreateBloombergExcelFile(self):
		file = join(getCurrentDirectory(), 'samples', 'sample PriceSTBF.xls')
		data = self.NavHandler.getSTBFNavDataFromFile(file)
		templateFile = join(getCurrentDirectory(), 'samples', 'Bloomberg fund pricing template.xlsx')
		outputDir = ""
		fundName = 'stbf'
		output_file_fullpath = self.NavHandler.createBloombergExcelFile(templateFile, outputDir, fundName, data)
		wb_obj = openpyxl.load_workbook(output_file_fullpath) 
		ws_obj=wb_obj.active
		first_data_row = 9
		#-- verify first row
		self.assertEqual(ws_obj.cell(first_data_row, 1).value, '12/30/2020')
		self.assertEqual(ws_obj.cell(first_data_row, 2).value, 'CLSTFBU HK Equity')
		self.assertEqual(ws_obj.cell(first_data_row, 3).value, 'CHINA LIFE FR ST BOND-B USD')
		self.assertEqual(ws_obj.cell(first_data_row, 4).value, 'USD')
		self.assertEqual(ws_obj.cell(first_data_row, 5).value, 9.9961)
		self.assertEqual(ws_obj.cell(first_data_row, 6).value, None)
		self.assertEqual(ws_obj.cell(first_data_row, 7).value, None)
		self.assertEqual(ws_obj.cell(first_data_row, 8).value, 48980171.47)
		self.assertEqual(ws_obj.cell(first_data_row, 9).value, 3998441.60)
		self.assertEqual(ws_obj.cell(first_data_row, 10).value, 400000)
		self.assertEqual(ws_obj.cell(first_data_row, 11).value, None)
		
		#-- verify second row
		self.assertEqual(ws_obj.cell(first_data_row+1, 1).value, '12/30/2020')
		self.assertEqual(ws_obj.cell(first_data_row+1, 2).value, 'CLSTFIU HK Equity')
		self.assertEqual(ws_obj.cell(first_data_row+1, 3).value, 'CHINA LIFE FR ST BOND-I USD')
		self.assertEqual(ws_obj.cell(first_data_row+1, 4).value, 'USD')
		self.assertEqual(ws_obj.cell(first_data_row+1, 5).value, 9.9959)
		self.assertEqual(ws_obj.cell(first_data_row+1, 6).value, None)
		self.assertEqual(ws_obj.cell(first_data_row+1, 7).value, None)
		self.assertEqual(ws_obj.cell(first_data_row+1, 8).value, 48980171.47)
		self.assertEqual(ws_obj.cell(first_data_row+1, 9).value, 44981729.87)
		self.assertEqual(ws_obj.cell(first_data_row+1, 10).value, 4500000)
		self.assertEqual(ws_obj.cell(first_data_row+1, 11).value, None)

	def testCreateBloombergExcelFileWith4FundColumn(self):
		file = join(getCurrentDirectory(), 'tests', 'testdata', 'PriceSTBF 2021-03-31editable.xls')
		data = self.NavHandler.getSTBFNavDataFromFile(file)
		templateFile = join(getCurrentDirectory(), 'samples', 'Bloomberg fund pricing template.xlsx')
		outputDir = ""
		fundName = 'stbf'
		output_file_fullpath = self.NavHandler.createBloombergExcelFile(templateFile, outputDir, fundName, data)
		wb_obj = openpyxl.load_workbook(output_file_fullpath) 
		ws_obj=wb_obj.active
		first_data_row = 9
		
		self.assertEqual(ws_obj.cell(first_data_row, 1).value, '03/31/2021')
		self.assertEqual(ws_obj.cell(first_data_row, 2).value, 'CLSTFBU HK Equity')
		self.assertEqual(ws_obj.cell(first_data_row, 3).value, 'CHINA LIFE FR ST BOND-B USD')
		self.assertEqual(ws_obj.cell(first_data_row, 4).value, 'USD')
		self.assertEqual(ws_obj.cell(first_data_row, 5).value, 10.0338)
		self.assertEqual(ws_obj.cell(first_data_row, 6).value, None)
		self.assertEqual(ws_obj.cell(first_data_row, 7).value, None)
		self.assertEqual(ws_obj.cell(first_data_row, 8).value, 249876982.57)
		self.assertEqual(ws_obj.cell(first_data_row, 9).value, 4013547.21)
		self.assertEqual(ws_obj.cell(first_data_row, 10).value, 400000)
		self.assertEqual(ws_obj.cell(first_data_row, 11).value, None)
		
		first_data_row = first_data_row + 1
		self.assertEqual(ws_obj.cell(first_data_row, 1).value, '03/31/2021')
		self.assertEqual(ws_obj.cell(first_data_row, 2).value, 'CLSTFIU HK Equity')
		self.assertEqual(ws_obj.cell(first_data_row, 3).value, 'CHINA LIFE FR ST BOND-I USD')
		self.assertEqual(ws_obj.cell(first_data_row, 4).value, 'USD')
		self.assertEqual(ws_obj.cell(first_data_row, 5).value, 10.0262)
		self.assertEqual(ws_obj.cell(first_data_row, 6).value, None)
		self.assertEqual(ws_obj.cell(first_data_row, 7).value, None)
		self.assertEqual(ws_obj.cell(first_data_row, 8).value, 249876982.57)
		self.assertEqual(ws_obj.cell(first_data_row, 9).value, 245860749.83)
		self.assertEqual(ws_obj.cell(first_data_row, 10).value, 24521823.7879)
		self.assertEqual(ws_obj.cell(first_data_row, 11).value, None)

		first_data_row = first_data_row + 1
		self.assertEqual(ws_obj.cell(first_data_row, 1).value, '03/31/2021')
		self.assertEqual(ws_obj.cell(first_data_row, 2).value, 'CLSTFAU HK Equity')
		self.assertEqual(ws_obj.cell(first_data_row, 3).value, 'CHINA LIFE FR ST BOND-A (USD) USD')
		self.assertEqual(ws_obj.cell(first_data_row, 4).value, 'USD')
		self.assertEqual(ws_obj.cell(first_data_row, 5).value, 10.0040)
		self.assertEqual(ws_obj.cell(first_data_row, 6).value, None)
		self.assertEqual(ws_obj.cell(first_data_row, 7).value, None)
		self.assertEqual(ws_obj.cell(first_data_row, 8).value, 249876982.57)
		self.assertEqual(ws_obj.cell(first_data_row, 9).value, 100.04)
		self.assertEqual(ws_obj.cell(first_data_row, 10).value, 10)
		self.assertEqual(ws_obj.cell(first_data_row, 11).value, None)
		
		first_data_row = first_data_row + 1
		self.assertEqual(ws_obj.cell(first_data_row, 1).value, '03/31/2021')
		self.assertEqual(ws_obj.cell(first_data_row, 2).value, 'CLSTFAH HK Equity')
		self.assertEqual(ws_obj.cell(first_data_row, 3).value, 'CHINA LIFE FR ST BOND-A (HKD) HKD')
		self.assertEqual(ws_obj.cell(first_data_row, 4).value, 'HKD')
		self.assertEqual(ws_obj.cell(first_data_row, 5).value, 10.0001)
		self.assertEqual(ws_obj.cell(first_data_row, 6).value, None)
		self.assertEqual(ws_obj.cell(first_data_row, 7).value, None)
		self.assertEqual(ws_obj.cell(first_data_row, 8).value, 249876982.57)
		self.assertEqual(ws_obj.cell(first_data_row, 9).value, 2585.49)
		self.assertEqual(ws_obj.cell(first_data_row, 10).value, 2010)
		self.assertEqual(ws_obj.cell(first_data_row, 11).value, None)


	def testCreateThomsonExcelFile(self):
		file = join(getCurrentDirectory(), 'samples', 'sample PriceSTBF.xls')
		data = self.NavHandler.getSTBFNavDataFromFile(file)
		templateFile = join(getCurrentDirectory(), 'samples', 'Thomson Reuters fund pricing template.xlsx')
		outputDir = ""
		fundName = 'stbf'
		output_file_fullpath = self.NavHandler.createThomsonExcelFile(templateFile, outputDir, fundName, data)
		wb_obj = openpyxl.load_workbook(output_file_fullpath) 
		ws_obj=wb_obj.active
		first_data_row = 2
		#-- verify first row
		self.assertEqual(ws_obj.cell(first_data_row, 1).value, 'HK0000664455')
		self.assertEqual(ws_obj.cell(first_data_row, 2).value, 'China Life Franklin Global-Short Term Bond B USD')
		self.assertEqual(ws_obj.cell(first_data_row, 3).value, 'USD')
		self.assertEqual(ws_obj.cell(first_data_row, 4).value, 9.9961)
		self.assertEqual(ws_obj.cell(first_data_row, 5).value, 48980171.47)
		self.assertEqual(ws_obj.cell(first_data_row, 6).value, 3998441.60)
		self.assertEqual(ws_obj.cell(first_data_row, 7).value, 400000)

		#-- verify second row
		self.assertEqual(ws_obj.cell(first_data_row+1, 1).value, 'HK0000664489')
		self.assertEqual(ws_obj.cell(first_data_row+1, 2).value, 'China Life Franklin Global-Short Term Bond I USD')
		self.assertEqual(ws_obj.cell(first_data_row+1, 3).value, 'USD')
		self.assertEqual(ws_obj.cell(first_data_row+1, 4).value, 9.9959)
		self.assertEqual(ws_obj.cell(first_data_row+1, 5).value, 48980171.47)
		self.assertEqual(ws_obj.cell(first_data_row+1, 6).value, 44981729.87)
		self.assertEqual(ws_obj.cell(first_data_row+1, 7).value, 4500000)


	def testCreateThomsonExcelFileWith4FundColumn(self):
		file = join(getCurrentDirectory(), 'tests', 'testdata', 'PriceSTBF 2021-03-31editable.xls')
		data = self.NavHandler.getSTBFNavDataFromFile(file)
		templateFile = join(getCurrentDirectory(), 'samples', 'Thomson Reuters fund pricing template.xlsx')
		outputDir = ""
		fundName = 'stbf'
		output_file_fullpath = self.NavHandler.createThomsonExcelFile(templateFile, outputDir, fundName, data)
		wb_obj = openpyxl.load_workbook(output_file_fullpath) 
		ws_obj=wb_obj.active
		data_row = 2
		#-- verify first row
		self.assertEqual(ws_obj.cell(data_row, 1).value, 'HK0000664455')
		self.assertEqual(ws_obj.cell(data_row, 2).value, 'China Life Franklin Global-Short Term Bond B USD')
		self.assertEqual(ws_obj.cell(data_row, 3).value, 'USD')
		self.assertEqual(ws_obj.cell(data_row, 4).value, 10.0338)
		self.assertEqual(ws_obj.cell(data_row, 5).value, 249876982.57)
		self.assertEqual(ws_obj.cell(data_row, 6).value, 4013547.21)
		self.assertEqual(ws_obj.cell(data_row, 7).value, 400000)

		data_row = data_row + 1
		self.assertEqual(ws_obj.cell(data_row, 1).value, 'HK0000664489')
		self.assertEqual(ws_obj.cell(data_row, 2).value, 'China Life Franklin Global-Short Term Bond I USD')
		self.assertEqual(ws_obj.cell(data_row, 3).value, 'USD')
		self.assertEqual(ws_obj.cell(data_row, 4).value, 10.0262)
		self.assertEqual(ws_obj.cell(data_row, 5).value, 249876982.57)
		self.assertEqual(ws_obj.cell(data_row, 6).value, 245860749.83)
		self.assertEqual(ws_obj.cell(data_row, 7).value, 24521823.7879)
		
		data_row = data_row + 1
		self.assertEqual(ws_obj.cell(data_row, 1).value, 'HK0000664422')
		self.assertEqual(ws_obj.cell(data_row, 2).value, 'China Life Franklin Global-Short Term Bond A (USD) USD')
		self.assertEqual(ws_obj.cell(data_row, 3).value, 'USD')
		self.assertEqual(ws_obj.cell(data_row, 4).value, 10.0040)
		self.assertEqual(ws_obj.cell(data_row, 5).value, 249876982.57)
		self.assertEqual(ws_obj.cell(data_row, 6).value, 100.04)
		self.assertEqual(ws_obj.cell(data_row, 7).value, 10)

		data_row = data_row + 1
		self.assertEqual(ws_obj.cell(data_row, 1).value, 'HK0000664430')
		self.assertEqual(ws_obj.cell(data_row, 2).value, 'China Life Franklin Global-Short Term Bond A (HKD) HKD')
		self.assertEqual(ws_obj.cell(data_row, 3).value, 'HKD')
		self.assertEqual(ws_obj.cell(data_row, 4).value, 10.0001)
		self.assertEqual(ws_obj.cell(data_row, 5).value, 249876982.57)
		self.assertEqual(ws_obj.cell(data_row, 6).value, 2585.49)
		self.assertEqual(ws_obj.cell(data_row, 7).value, 2010)

	# def testUpdateWebSite(self):
	# 	#-- setup
	# 	file = join(getCurrentDirectory(), 'samples', 'sample PriceSTBF.xls')
	# 	mode = Constants.MODE_TEST
	# 	fundName = "stbf"
	# 	timeOut = 10
	# 	#-- loop the iterator and send request to website
	# 	all_nav = list(NavHandler().getSTBFNavDataFromFile(file))
	# 	for i in range(len(all_nav)):
	# 		NavHandler().updateWebSite(mode, timeOut, fundName, all_nav[i])